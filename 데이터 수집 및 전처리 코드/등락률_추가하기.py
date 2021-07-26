'''
http://konex.krx.co.kr/contents/MKD/04/0402/04020100/MKD04020100T3.jsp
에서 주가를 엑셀파일로 받아와야지만 사용 가능하다.
'''

import openpyxl as op

last = input('엑셀의 가장 마지막 행의 숫자를 입력해주세요:')
filepath = "data.xlsx"

wb = op.load_workbook(filepath) #workbook 객체 생성 

ws = wb.active  #worksheet 객체 생성 

for r in ws.rows: # ws.rows는 가장 마지막 행이 있는 data의 위치를 나타낸다.
    row_index = r[0].row #r[0]는 for문의 각 행값에 따른 열 위치를 의미한다. 숫자 0은 첫번째 열을 말한다.
    ws['K1'] = '등락률'
    ws['K'+str(row_index)] =  "=C"+str(row_index)+"/B"+str(row_index+1)+"*100" #등락률 구하는 공식
    ws['K'+last] = '0.0' #마지막 값은 계산이 안돼서 임의로 값을 집어 넣음
    wb.save("data1.xlsx")
    
### data1.xlsx파일을 열어서 직접 저장을 누른 후 사용 가능하다.
